import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# ─── CONFIGURATION ────────────────────────────────────────────────────────────
ROOT_FOLDER  = r"E:\Apps Projects\2. Mehandi Magic\1 New Download Designs mix\Grouped Design for JSON"   # <-- change this path
OUTPUT_FILE  = r"E:\Apps Projects\2. Mehandi Magic\1 New Download Designs mix\name of all designs.xlsx"  # <-- output Excel path
# ──────────────────────────────────────────────────────────────────────────────

SUPPORTED_EXTENSIONS = {".jpg", ".jpeg", ".png", ".webp", ".bmp", ".gif", ".tiff", ".tif"}

def collect_images(root):
    images = []
    for dirpath, dirnames, filenames in os.walk(root):
        dirnames.sort()
        for filename in sorted(filenames):
            ext = os.path.splitext(filename)[1].lower()
            if ext in SUPPORTED_EXTENSIONS:
                rel_folder = os.path.relpath(dirpath, root)
                images.append({
                    "sr":       0,
                    "name":     filename,
                    "folder":   rel_folder if rel_folder != "." else "(root)",
                    "ext":      ext.lstrip(".").upper(),
                    "fullpath": os.path.join(dirpath, filename),
                })
    for i, img in enumerate(images, start=1):
        img["sr"] = i
    return images

def border_all(thin):
    return Border(left=thin, right=thin, top=thin, bottom=thin)

def create_excel(images, output_path):
    wb = Workbook()
    ws = wb.active
    ws.title = "Image Names"

    # ── styles ──
    thin = Side(style="thin", color="CCCCCC")
    bd   = border_all(thin)

    header_font  = Font(name="Arial", bold=True, color="FFFFFF", size=11)
    header_fill  = PatternFill("solid", start_color="3B5998")
    header_align = Alignment(horizontal="center", vertical="center")

    alt_fill  = PatternFill("solid", start_color="EEF2FF")
    center    = Alignment(horizontal="center", vertical="center")
    left      = Alignment(horizontal="left",   vertical="center")

    # ── headers ──
    headers = ["Sr. No.", "Image Name", "Folder", "Extension"]
    col_widths = [10, 40, 35, 12]
    for col, (h, w) in enumerate(zip(headers, col_widths), start=1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font      = header_font
        cell.fill      = header_fill
        cell.alignment = header_align
        cell.border    = bd
        ws.column_dimensions[cell.column_letter].width = w

    ws.row_dimensions[1].height = 22

    # ── data rows ──
    for row_idx, img in enumerate(images, start=2):
        values = [img["sr"], img["name"], img["folder"], img["ext"]]
        fill = alt_fill if row_idx % 2 == 0 else None

        for col, val in enumerate(values, start=1):
            cell = ws.cell(row=row_idx, column=col, value=val)
            cell.border    = bd
            cell.alignment = center if col in (1, 4) else left
            cell.font      = Font(name="Arial", size=10)
            if fill:
                cell.fill = fill

        ws.row_dimensions[row_idx].height = 18

    # ── summary row ──
    total_row = len(images) + 2
    ws.cell(row=total_row, column=1, value="Total").font = Font(name="Arial", bold=True, size=10)
    ws.cell(row=total_row, column=2, value=len(images)).font = Font(name="Arial", bold=True, size=10)
    for col in range(1, 5):
        ws.cell(row=total_row, column=col).border = bd
        ws.cell(row=total_row, column=col).fill   = PatternFill("solid", start_color="D9E1F2")

    # ── freeze header ──
    ws.freeze_panes = "A2"

    wb.save(output_path)
    print(f"Saved: {output_path}  ({len(images)} images)")

if __name__ == "__main__":
    import sys
    if not os.path.isdir(ROOT_FOLDER):
        print(f"ERROR: ROOT_FOLDER not found:\n  {ROOT_FOLDER}")
        sys.exit(1)

    images = collect_images(ROOT_FOLDER)
    if not images:
        print("No images found.")
        sys.exit(0)

    create_excel(images, OUTPUT_FILE)
